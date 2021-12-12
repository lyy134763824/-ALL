from openpyxl import load_workbook

wb = load_workbook('data1.xlsx')
sheets = wb.worksheets  # 获取当前所有的sheet
print(sheets)
'''
如果碰到带有公式的表格，而只想要读取其中计算的结果时，可以在读取工作簿的时候加上data_only=True的属性，例如：
from openpyxl import load_workbook
wb = load_workbook('data1.xlsx', data_only=True)

'''
# 获取第一张sheet
sheet1 = sheets[0]
# sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
print(sheet1)

# 通过Cell对象读取
cell_11 = sheet1.cell(1, 1).value
print(cell_11)
cell_11 = sheet1.cell(1, 2).value
print(cell_11)
# 获取第一行所有数据
row1 = []
print(sheet1[1])
for row in sheet1[1]:
    print(row)
    row1.append(row.value)
print(row1)

# 获取第一列所有数据
col1 = []
for col in sheet1['A']:
    col1.append(col.value)
print(col1)
rows = sheet1.rows
columns = sheet1.columns
print(rows)
print(columns)
print('')

# 迭代读取所有的行
for row in rows:
    print(row)
    row_val = [col.value for col in row]
    print(row_val)
print('')

# 迭代读取所有的列
for col in columns:
    print(col)
    col_val = [row.value for row in col]




